import allure
import config
from TestVehicle.pom.vehicleLocator import VehicleLocator
from random import choices, choice
from openpyxl import load_workbook


class VehicleImportFixture:

    def __init__(self):
        self.driver = None

    def vehicle_open(self):
        """Открытие формЫ импорта ТС"""
        vehicle = VehicleLocator(self.driver)
        with allure.step("Загрузка страницы ТС"):
            assert vehicle.get_button("Добавить ТС")
        with allure.step("Переход в импорт"):
            vehicle.get_button("Загрузить ТС из файла").click()
        with allure.step("Загрузки страницы импорта ТС"):
            assert vehicle.get_button("Закрыть")

    def vehicle_import(self):
        """Раздел импорта ТС"""
        vehicle = VehicleLocator(self.driver)
        with allure.step("Выбор шаблона"):
            if vehicle.get_template().text != "Шаблон ТС (стандартный)":
                vehicle.get_template().click()
                vehicle.get_drop_down_find().send_keys("Шаблон ТС (стандартный)")
                vehicle.get_drop_down_find().send_keys('\n')
                vehicle.get_drop_down_meaning("Шаблон ТС (стандартный)").click()
        with allure.step("загрузка файла"):
            vehicle.get_upload().send_keys(config.url_file)
        with allure.step('Ожидание загрузки файла'):
            assert vehicle.get_button("Загрузить")

    def vehicle_owner(self):
        """Раздел "Страхователь"""
        vehicle = VehicleLocator(self.driver)
        with allure.step("Компания-страхователь"):
            vehicle.get_drop_down("Компания-страхователь").click()
            vehicle.get_drop_down_meaning(config.vehicle_dict["Компания"]).click()

    def vehicle_file(self):
        """Формирование файла со случайными параметрами"""
        with allure.step("Формируем файл"):
            wb = load_workbook(config.url_file)
            sheet = wb.active
            sheet['B4'] = config.vehicle_dict["ИНН_Моэск"]
            sheet['C4'] = config.vehicle_dict["КПП_Моэск"]
            key = ['I4', 'J4', 'K4', 'L4']
            for i in key:
                sheet[i] = ""
            key_parameter = choice(key)
            if key_parameter == 'L4':
                sheet[key_parameter] = "".join((choices(config.str_rus, k=1)) + (choices(config.str_number, k=3)) +
                                               (choices(config.str_rus, k=2)) + (choices(config.str_number, k=3)))
            else:
                sheet[key_parameter] = "".join(choices(config.str_vin, k=17))
            wb.save(config.url_file)
            wb.close()

    def vehicle_file_addition(self):
        """Формирование файла с параметрами внешней системы (для проверки автодополнения)"""
        with allure.step("Формируем файл"):
            wb = load_workbook(config.url_file_addition)
            sheet = wb.active
            for i in range(3, 14):
                key = sheet.cell(row=4, column=i)
                sheet[key] = ""
        with allure.step("VIN"):
            sheet['I4'] = config.vehicle_dict["VIN"]

    def vehicle_close(self):
        """Закрытие формы добавления ТС"""
        vehicle = VehicleLocator(self.driver)
        with allure.step('Загрузить'):
            vehicle.get_button("Загрузить").click()
        with allure.step("Загрузки промежуточной таблицы"):
            vehicle.get_not_button("Загрузить")
            assert vehicle.get_button("К списку ТС")
        with allure.step("Проверка добавления ТС"):
            assert vehicle.get_upload_stat() == vehicle.UPLOAD_STAT
